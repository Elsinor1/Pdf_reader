import os
import tkinter
import sqlite3
import threading
import openpyxl
import customtkinter as CTk
from tkinter import messagebox
from helpers import (
    get_folder_path,
    get_file_path,
    create_db,
    read_pdf,
    search_keyword,
    search_between,
    ocr_pdf,
)
from PIL import Image

CTk.set_appearance_mode("Dark")  # Modes: "System" (standard), "Dark", "Light"
CTk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"


class App(CTk.CTk):
    def __init__(self):
        super().__init__()

        # WINDOW CONFIGURATION
        self.title("Machine reader")
        self.geometry(f"{1400}x{600}")

        # layout grid 4x4
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(2, weight=2, minsize=600)
        self.grid_columnconfigure(1, weight=2, minsize=600)

        # ---HEADER----
        self.header_frame = CTk.CTkFrame(self, corner_radius=0)
        self.header_frame.grid(row=0, column=0, sticky="nsew", pady=(20, 20))
        self.header_label = CTk.CTkLabel(
            self.header_frame,
            text="Machine\nReader",
            font=CTk.CTkFont(size=30, weight="bold"),
        )
        self.header_frame.grid_columnconfigure(0, weight=1)
        self.header_label.grid(row=1, column=0, sticky="ew")

        # Title image
        bg = CTk.CTkImage(
            light_image=Image.open("images/AI_reader.png"),
            dark_image=Image.open("images/AI_reader.png"),
            size=(190, 250),
        )

        bg_label = CTk.CTkLabel(
            self.header_frame,
            text="",
            image=bg,
        )
        bg_label.grid(row=0, column=0, padx=20, pady=(20, 20), sticky="ewsn")
        # bg_button.place(x=0, y=0, relwidth=0.2, relheight=0.2)

        # ---SIDEBAR----
        # Frame
        self.leftbar_frame = CTk.CTkFrame(self, width=50, corner_radius=0)
        self.leftbar_frame.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        self.leftbar_frame.grid_columnconfigure((0, 2), weight=2)
        self.leftbar_frame.grid_columnconfigure(1, weight=1)
        self.leftbar_frame.grid_columnconfigure((1, 2), minsize=200)

        # Read PDF document label
        self.label_1 = CTk.CTkLabel(
            self.leftbar_frame,
            text="Read PDF document",
            font=CTk.CTkFont(size=20, weight="bold"),
        )
        self.label_1.grid(
            row=0, column=0, columnspan=3, padx=20, pady=(20, 0), sticky="we"
        )

        # Select input
        self.label_in = CTk.CTkLabel(
            self.leftbar_frame,
            text="Select input",
            font=CTk.CTkFont(size=18, weight="bold"),
        )
        self.label_in.grid(
            row=1, column=0, columnspan=3, padx=20, pady=(20, 5), sticky="we"
        )

        # Choose folder Button
        self.leftbar_button_1 = CTk.CTkButton(
            self.leftbar_frame,
            command=self.load_folder_input1,
            text="Select input folder",
        )
        self.leftbar_button_1.grid(row=2, column=0, padx=(20, 5), pady=5)

        # Amount of files text field
        self.label_amount = CTk.CTkLabel(
            self.leftbar_frame,
            text="Amount of files",
            font=CTk.CTkFont(size=15, weight="bold"),
        )
        self.label_amount.grid(row=2, column=1, padx=5, pady=5, sticky="we")
        self.leftbar_amount = CTk.CTkTextbox(
            self.leftbar_frame, width=100, height=15, font=CTk.CTkFont(size=15)
        )
        self.leftbar_amount.grid(row=2, column=2, padx=(5, 20), pady=5)

        # Path text field
        self.leftbar_path_1 = CTk.CTkTextbox(
            self.leftbar_frame, width=400, height=15, font=CTk.CTkFont(size=12)
        )
        self.leftbar_path_1.grid(
            row=4, column=0, columnspan=3, padx=20, pady=(5, 20), sticky="ew"
        )

        # Select output
        self.label_2 = CTk.CTkLabel(
            self.leftbar_frame,
            text="Select output",
            font=CTk.CTkFont(size=18, weight="bold"),
        )
        self.label_2.grid(
            row=5, column=0, columnspan=3, padx=(0, 5), pady=(20, 10), sticky="we"
        )
        # Radio button
        self.radiobutton_frame = CTk.CTkFrame(self.leftbar_frame)
        self.radiobutton_frame.grid(
            row=6, column=0, columnspan=2, padx=(20, 5), pady=(5, 5), sticky="nsew"
        )
        self.radiobutton_frame.grid_columnconfigure(0, weight=3)
        self.radiobutton_frame.grid_columnconfigure(1, weight=4)
        self.radio_out_1 = tkinter.IntVar(value=0)
        self.radio_button_1 = CTk.CTkRadioButton(
            self.radiobutton_frame,
            variable=self.radio_out_1,
            text="Choose existing db",
            value=0,
        )
        self.radio_button_1.grid(row=0, column=0, pady=10, padx=5, sticky="n")
        self.radio_button_2 = CTk.CTkRadioButton(
            self.radiobutton_frame,
            variable=self.radio_out_1,
            text="Create new db",
            value=1,
        )
        self.radio_button_2.grid(row=0, column=1, pady=10, padx=5, sticky="n")
        # Select output folder
        self.leftbar_button_2 = CTk.CTkButton(
            self.leftbar_frame, command=self.load_output_1, text="Select db or folder"
        )
        self.leftbar_button_2.grid(row=6, column=2, padx=(5, 20), pady=5)
        # Path for output
        self.leftbar_path_2 = CTk.CTkTextbox(
            self.leftbar_frame, width=400, height=15, font=CTk.CTkFont(size=12)
        )
        self.leftbar_path_2.grid(
            row=7, column=0, columnspan=3, padx=(20, 20), pady=(5, 20), sticky="WESN"
        )
        # Read button
        self.leftbar_button_read = CTk.CTkButton(
            self.leftbar_frame,
            text="READ",
            command=lambda: threading.Thread(target=self.read_write).start(),
        )
        self.leftbar_button_read.grid(
            row=8, column=0, padx=(20, 5), pady=(5, 20), sticky="WESN"
        )
        # Update radio button
        self.radiobutton_frame_2 = CTk.CTkFrame(
            self.leftbar_frame,
        )
        self.radiobutton_frame_2.grid(
            row=8, column=1, columnspan=2, padx=(5, 20), pady=(5, 20), sticky="nsew"
        )
        # self.radiobutton_frame_2.grid_columnconfigure(0, weight=3)
        # self.radiobutton_frame_2.grid_columnconfigure(1, weight=4)
        self.radio_out_2 = tkinter.IntVar(value=0)
        self.radio_button_3 = CTk.CTkRadioButton(
            self.radiobutton_frame_2,
            variable=self.radio_out_2,
            text="Update existing data",
            value=0,
        )
        self.radio_button_3.grid(row=0, column=0, pady=5, padx=5, sticky="nsew")
        self.radio_button_4 = CTk.CTkRadioButton(
            self.radiobutton_frame_2,
            variable=self.radio_out_2,
            text="Don't update existing data",
            value=1,
        )
        self.radio_button_4.grid(row=0, column=1, pady=5, padx=5, sticky="nsew")
        # Left progress bar
        self.progressbar_1 = CTk.CTkProgressBar(
            self.leftbar_frame, height=20, corner_radius=5, border_width=3
        )
        # Ocr usage checkbox
        self.ocr_checkbox = CTk.CTkCheckBox(
            master=self.radiobutton_frame_2,
            text="Use OCR to read pdf's",
            onvalue=True,
            offvalue=False,
        )
        self.ocr_checkbox.grid(
            row=1, column=0, columnspan=2, sticky="wesn", padx=(5, 20), pady=(5, 5)
        )
        self.ocr_checkbox.select()

        # Progressbar text box
        self.progressbar_1_text = CTk.CTkLabel(
            self.radiobutton_frame_2,
            text="",
            fg_color="transparent",
            width=50,
            font=CTk.CTkFont(size=15, weight="bold"),
        )

        # -----RIGHT SECTION----
        # Frame
        self.rightbar_frame = CTk.CTkFrame(self)
        self.rightbar_frame.grid(row=0, column=2, padx=(0, 20), pady=20, sticky="nswe")
        self.rightbar_frame.grid_columnconfigure((0), weight=0)
        self.rightbar_frame.grid_columnconfigure((1), weight=3)
        # Label
        self.label_right = CTk.CTkLabel(
            self.rightbar_frame,
            text="Analyze",
            font=CTk.CTkFont(size=20, weight="bold"),
        )
        self.label_right.grid(
            row=0, column=0, columnspan=2, padx=20, pady=(20, 20), sticky="we"
        )
        # Choose input file Button
        self.right_button_1 = CTk.CTkButton(
            self.rightbar_frame,
            text="Select DB to analyze",
            command=self.load_input_2,
        )
        self.right_button_1.grid(
            row=1, column=0, padx=(20, 5), sticky="we", pady=(5, 5)
        )
        # Path text field
        self.right_path_1 = CTk.CTkTextbox(
            self.rightbar_frame, width=50, height=15, font=CTk.CTkFont(size=12)
        )
        self.right_path_1.grid(row=1, column=1, padx=(5, 20), sticky="we", pady=(5, 5))

        # What to analyze
        self.label_analyze = CTk.CTkLabel(
            self.rightbar_frame,
            text="Select search criteria",
            font=CTk.CTkFont(size=15, weight="bold"),
        )
        self.label_analyze.grid(
            row=2, column=0, columnspan=2, padx=20, pady=(15, 5), sticky="we"
        )

        # analyze input radio button
        self.input_slider_frame = CTk.CTkFrame(self.rightbar_frame)
        self.input_slider_frame.grid(
            row=3, column=0, columnspan=2, sticky="ew", pady=(5, 5)
        )
        self.input_slider_frame.grid_columnconfigure((0, 1), weight=1)
        self.input_slider_frame.grid_rowconfigure((0, 1), weight=2)
        self.input_slider_frame.grid_rowconfigure((3, 4), weight=1)

        # input variables
        self.analyze_input_type = tkinter.IntVar(value=0)
        # Input buttons
        self.input_1 = CTk.CTkRadioButton(
            master=self.input_slider_frame,
            text="Single word",
            command=self.single_or_double,
            variable=self.analyze_input_type,
            value=0,
        )
        self.input_1.grid(row=0, column=0, pady=(20, 5), padx=(20, 5), sticky="w")
        self.input_2 = CTk.CTkRadioButton(
            master=self.input_slider_frame,
            text="Search between",
            command=self.single_or_double,
            variable=self.analyze_input_type,
            value=1,
        )
        self.input_2.grid(row=1, column=0, pady=(5, 5), padx=(20, 5), sticky="w")

        # Search input fields
        self.analyze_input_1 = CTk.CTkEntry(
            self.input_slider_frame,
            placeholder_text="Keyword",
            font=CTk.CTkFont(size=12),
        )
        self.analyze_input_1.grid(
            row=3, column=0, padx=(20, 5), pady=(5, 5), sticky="nw"
        )
        self.analyze_input_2 = CTk.CTkEntry(
            self.input_slider_frame,
            state="disabled",
            placeholder_text="",
            font=CTk.CTkFont(size=12),
        )
        self.analyze_input_2.grid(
            row=4, column=0, padx=(20, 5), pady=(5, 20), sticky="nw"
        )

        # Add or delete file
        self.button_frame = CTk.CTkFrame(self.input_slider_frame)
        self.button_frame.grid(row=0, column=1, sticky="ew", padx=20, pady=(20, 5))
        self.button_frame.grid_columnconfigure((0, 1), weight=1)
        self.right_button_1 = CTk.CTkButton(
            self.button_frame, text="Add to list", command=self.add_to_list
        )
        self.right_button_1.grid(row=0, column=0, padx=10, sticky="we", pady=(5, 5))
        self.right_button_1 = CTk.CTkButton(
            self.button_frame, text="Remove from list", command=self.remove_from_list
        )
        self.right_button_1.grid(row=0, column=1, padx=10, sticky="we", pady=(5, 5))
        # Special search list box
        self.list_count = tkinter.IntVar(value=0)
        self.analyze_list = tkinter.Listbox(
            self.input_slider_frame, font=CTk.CTkFont(size=12)
        )
        self.analyze_list.grid(
            row=1, column=1, rowspan=4, padx=20, pady=(5, 20), sticky="news"
        )

        # Select output
        self.label2 = CTk.CTkLabel(
            self.rightbar_frame,
            text="Select output",
            font=CTk.CTkFont(size=15, weight="bold"),
        )
        self.label2.grid(
            row=4, column=0, columnspan=2, padx=20, pady=(0, 5), sticky="we"
        )
        # Select output folder
        self.right_button_2 = CTk.CTkButton(
            self.rightbar_frame, text="Select output folder", command=self.load_output_2
        )
        self.right_button_2.grid(
            row=5, column=0, padx=(20, 5), pady=(5, 5), sticky="ew"
        )
        # Path for output
        self.right_path_2 = CTk.CTkTextbox(
            self.rightbar_frame, height=15, font=CTk.CTkFont(size=12)
        )
        self.right_path_2.grid(row=5, column=1, sticky="we", padx=(5, 20), pady=(5, 5))
        self.progressbar_2_text = CTk.CTkLabel(self.rightbar_frame)

        # Analyze button
        self.right_button_read = CTk.CTkButton(
            self.rightbar_frame,
            command=lambda: threading.Thread(target=self.analyze).start(),
            text="ANALYZE",
            height=50,
        )
        self.right_button_read.grid(
            row=6,
            column=0,
            padx=(20, 5),
            pady=(5, 20),
            sticky="WESN",
        )
        # Right progress bar
        progressbar_2_var = tkinter.IntVar(value=0)
        self.progressbar_2 = CTk.CTkProgressBar(self.rightbar_frame)

    # Displays single or double entry fields
    def single_or_double(self):
        if self.analyze_input_type.get() == 0:
            self.analyze_input_1.configure(placeholder_text="Keyword")
            self.analyze_input_2.configure(state="disabled", placeholder_text="")

        else:
            self.analyze_input_1.configure(placeholder_text="Between this..")
            self.analyze_input_2.configure(
                state="normal", placeholder_text=" ..and this"
            )
        return

    def load_folder_input1(self):
        path_in_1 = get_folder_path()
        if path_in_1:
            # Writes amount of files to leftbar_amount text field
            self.leftbar_amount.delete("0.0", "end")
            self.leftbar_amount.insert("0.0", len(os.listdir(path_in_1)))

            # Writes to leftbar_path1 field
            self.leftbar_path_1.delete("0.0", "end")
            self.leftbar_path_1.insert("0.0", path_in_1)
        return

    def load_output_1(self):
        # If radio button chosen existing db
        if self.radio_out_1.get() == 0:
            path_out_1 = get_file_path()
        else:
            path_out_1 = get_folder_path()

        # Writes to leftbar_path2 field
        self.leftbar_path_2.delete("0.0", "end")
        self.leftbar_path_2.insert("0.0", path_out_1)
        return

    def load_output_2(self):
        path_out_2 = get_folder_path()
        self.right_path_2.delete("0.0", "end")
        self.right_path_2.insert("0.0", path_out_2)
        return

    def load_input_2(self):
        path = get_file_path()
        self.right_path_1.delete("0.0", "end")
        self.right_path_1.insert("0.0", path)
        return

    def add_to_list(self):
        if self.analyze_input_type.get() == 0:
            self.analyze_list.insert(
                self.list_count.get(), ("KEYWORD:", self.analyze_input_1.get())
            )
        else:
            self.analyze_list.insert(
                self.list_count.get(),
                ("BETWEEN:", self.analyze_input_1.get(), self.analyze_input_2.get()),
            )
        self.list_count.set(value=self.list_count.get() + 1)
        return

    def remove_from_list(self):
        self.analyze_list.delete(self.analyze_list.curselection())
        return

    def read_write(self):
        # DB connection setup
        if self.radio_out_1.get() == 0:  # If user chose an output db
            output_db_path = self.leftbar_path_2.get("0.0", "end-1c")
        else:  # If user wants to create new output db
            output_db_path = create_db(self.leftbar_path_2.get("0.0", "end-1c"))
        try:
            db = sqlite3.connect(output_db_path)
            cur = db.cursor()
            cur.execute(
                "CREATE TABLE IF NOT EXISTS pdf_data(pdf_name TEXT UNIQUE NOT NULL PRIMARY KEY, pdf_text TEXT)"
            )
        except sqlite3.DatabaseError:
            messagebox.showerror(message=f"Selected database is invalid")
            return
        except TypeError:
            return

        # Gets list of PNs that already exist in the DB
        file_name_tuples = cur.execute("SELECT pdf_name FROM pdf_data").fetchall()
        existing_file_names = []
        for file_name in file_name_tuples:
            existing_file_names.append(file_name[0])

        # Reading
        global input_path
        global text_list
        input_path = self.leftbar_path_1.get("0.0", "end-1c")
        text_list = []
        try:
            files = os.listdir(input_path)
        except FileNotFoundError:
            messagebox.showerror(message="Invalid input path")
            return
        # Setup for progressbar before reading
        self.progressbar_1.set(0)
        self.progressbar_1.grid(
            row=9, column=0, columnspan=3, padx=(20, 20), pady=(5, 20), sticky="nsew"
        )
        # self.radiobutton_frame_2.grid_forget()
        self.radio_button_3.grid_forget()
        self.radio_button_4.grid_forget()
        self.progressbar_1_text.grid(
            row=8, column=1, columnspan=2, padx=(5, 20), pady=(5, 5), sticky="nsew"
        )
        files_total = len(files)

        for index, file in enumerate(files):
            # Display progress
            self.progressbar_1.set(index / files_total)
            self.progressbar_1_text.configure(
                text=f"READING IN PROGRESS: {index} / {files_total}"
            )
            self.update_idletasks
            # If update is not wanted skip to next file
            if self.radio_out_2 == 1 and file in existing_file_names:
                continue

            # Creates path to a file
            file_path = input_path + "/" + file
            # Reads text from PDF
            text = read_pdf(file_path)
            # If read pdf does not work
            if (
                text == ""
                or text == "Cannot be read"
                and self.ocr_checkbox.get() == True
            ):
                text = ocr_pdf(file_path, input_path)

            # Gets the part number from file name by striping .pdf
            pn_split = file.split(".")
            pn = pn_split[0]
            # Outputs a list of dictionaries pn : text
            text_list.append({"pn": pn, "pdf_text": text})

        # WRITING
        # Setup for progressbar before writing to DB
        self.progressbar_1.set(0)
        files_total = len(text_list)
        # List of not successful writing to DB
        not_successful = []
        # For every read drawing
        for index, line in enumerate(text_list, start=1):
            # If PN already exists in the database, update drawing text
            if line["pn"] in existing_file_names:
                cur.execute(
                    "UPDATE pdf_data SET pdf_text = ? WHERE pdf_name = ?",
                    (line["pdf_text"], line["pn"]),
                )
            # If PN is new
            else:
                try:
                    cur.execute(
                        "INSERT INTO pdf_data (pdf_name, pdf_text) VALUES(?, ?)",
                        (line["pn"], line["pdf_text"]),
                    )
                except:
                    not_successful.append(line["pn"])

            self.progressbar_1.set(index / files_total)
            self.progressbar_1_text.configure(
                text=f"SAVING TO DB: {index} / {files_total}"
            )
            self.update_idletasks

        if len(not_successful) > 0:
            messagebox.showinfo(
                message=f"Read is finished, {len(not_successful)} failed"
            )
        else:
            messagebox.showinfo(message="Read is finished")

        self.progressbar_1_text.configure(text=f"")
        self.progressbar_1.grid_forget()
        self.progressbar_1_text.grid_forget()
        self.radio_button_3.grid(row=0, column=0, pady=10, padx=5, sticky="n")
        self.radio_button_4.grid(row=0, column=1, pady=10, padx=5, sticky="n")
        # self.radiobutton_frame_2.grid(
        #     row=8, column=1, columnspan=2, padx=(5, 20), pady=(5, 5), sticky="nsew"
        # )
        db.commit()
        db.close()
        return

    def analyze(self):
        # Checks if the output excel sheet can be created
        if self.right_path_2.get("0.0", "end-1c") == "":
            messagebox.showerror(message="Output folder was not selected")
            return
        output_path = self.right_path_2.get("0.0", "end-1c") + "/extract.xlsx"
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Export"
            wb.save(output_path)
        except PermissionError:
            messagebox.showerror(message="Failed to create output file")
            return
        # Connects to a DB
        db_path = self.right_path_1.get("0.0", "end-1c")
        try:
            db = sqlite3.connect(db_path)
            cur = db.cursor()
            # Reads the DB
            input_data = cur.execute("SELECT * FROM pdf_data").fetchall()
        except sqlite3.OperationalError:
            messagebox.showerror(message="DB table not found")
            return
        # Progressbar setup
        self.progressbar_2.grid(
            row=6, column=1, padx=(5, 20), pady=(5, 20), sticky="nsew"
        )
        self.progressbar_2_text.grid(row=5, column=1, padx=(5, 20), pady=(5, 5))
        self.right_path_2.grid_forget()
        # Reads search criteria
        analyze_parametres = self.analyze_list.get("0", self.analyze_list.size() - 1)
        output = []
        input_total = len(input_data)
        # For every line in a DB
        for index, line in enumerate(input_data, start=1):
            file_name = line[0]
            test_analyzed = []
            # for every parameter:
            for parameter in analyze_parametres:
                # Search for keyword
                if parameter[0] == "KEYWORD:":
                    test_analyzed.append(search_keyword(parameter[1], line[1]))
                # Search between
                elif parameter[0] == "BETWEEN:":
                    test_analyzed.append(search_between(parameter, line[1]))
            output.append((file_name, test_analyzed))
            # Progressbar
            self.progressbar_2.set(index / input_total)
            self.progressbar_2_text.configure(
                text=f"ANALYZING DB: {index} / {input_total}"
            )
            self.update_idletasks

        # Writing to excel sheet
        # Headers
        ws.cell(row=1, column=1).value = "File name"
        for index, parameter in enumerate(analyze_parametres):
            if parameter[0] == "KEYWORD:":
                ws.cell(row=1, column=index + 2).value = parameter[1]
            else:
                ws.cell(
                    row=1, column=index + 2
                ).value = f"{parameter[1]} {parameter[2]}"

        # Output writing output=[(pdf_name,[param1, param2...]), (..)]
        for row_index, row in enumerate(output, start=2):
            ws.cell(row=row_index, column=1).value = row[0]
            data_columns = row[1]
            for col_index, column in enumerate(data_columns, start=2):
                try:
                    ws.cell(row=row_index, column=col_index).value = column
                except:
                    normal_string = "".join(ch for ch in column if ch.isalnum())
                    ws.cell(row=row_index, column=col_index).value = normal_string

            self.progressbar_2_text.configure(
                text=f"SAVING TO EXCEL FILE: {row_index - 1} / {input_total}"
            )
            self.progressbar_2.set(row_index / input_total)
            self.update_idletasks()

        wb.save(output_path)

        messagebox.showinfo(message="Analyze is finished")
        self.progressbar_2_text.configure(text=f"")
        self.progressbar_2.grid_forget()
        self.progressbar_2_text.grid_forget()
        self.right_path_2.grid(row=5, column=1, padx=(5, 20), pady=(5, 5), sticky="ew")

        return


if __name__ == "__main__":
    app = App()
    app.mainloop()
